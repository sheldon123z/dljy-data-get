#!/usr/bin/env python3
"""按「地区 / 月 / 周」三级目录导出 Excel，支持增量重建。

    data/exports/tree/
    ├── 新疆/
    │   ├── 2026-01/
    │   │   ├── 新疆_2026-01_W1_0128-0131.xlsx
    │   │   └── 新疆_2026-01_月汇总.xlsx
    │   ├── 2026-02/…
    │   └── 新疆_全期汇总.xlsx
    └── _索引.xlsx

周的切法默认按自然周（周一起算），也可以用 --week-mode month
按「当月第几个 7 天」切（7-1 = 1–7 日，7-2 = 8–14 日……）。

增量策略：每个周文件记录一份内容指纹，只有数据真的变了才重写，
新增的周直接补上，被重采覆盖的周会被替换。--full 可强制全部重写。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

import config  # noqa: E402
from common import (  # noqa: E402
    DETAIL_FILE,
    METADATA_FILE,
    PRICE_UNIT,
    as_number,
    iter_csv,
    load_provinces,
    mean,
    now_stamp,
    to_date,
)

HEAD_FILL = "1F3B63"
SUB_FILL = "DCE4F0"
WHITE = "FFFFFF"

DETAIL_HEADERS = [
    ("trade_date", "交易日期", 13),
    ("weekday", "星期", 7),
    ("time_slot", "时点", 9),
    ("day_ahead_price", "日前价格", 12),
    ("real_time_price", "实时价格", 12),
    ("spread", "实时-日前", 12),
]
DAILY_HEADERS = [
    ("trade_date", "交易日期", 13),
    ("weekday", "星期", 7),
    ("point_count", "分时点数", 10),
    ("day_ahead_avg", "日前均价", 12),
    ("real_time_avg", "实时均价", 12),
    ("spread_avg", "实时-日前", 12),
    ("day_ahead_min", "日前最低", 12),
    ("day_ahead_max", "日前最高", 12),
    ("real_time_min", "实时最低", 12),
    ("real_time_max", "实时最高", 12),
    ("negative_points", "负价点", 9),
    ("zero_points", "零价点", 9),
]
WEEKDAYS = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

STATE_FILE = "_增量状态.json"


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="按地区/月/周三级目录导出 Excel")
    parser.add_argument("--raw-dir", type=Path, default=config.RAW_DIR)
    parser.add_argument("--output-dir", type=Path, default=config.EXPORT_DIR / "tree")
    parser.add_argument(
        "--week-mode",
        choices=["iso", "month"],
        default="month",
        help="iso=自然周（周一起算）；month=当月第几个 7 天（默认，对应 7-1/7-2 这种叫法）",
    )
    parser.add_argument("--only-provinces", nargs="*", help="只导出指定区域")
    parser.add_argument("--full", action="store_true", help="忽略增量状态，全部重写")
    parser.add_argument("--no-summary", action="store_true", help="跳过月汇总与全期汇总")
    return parser.parse_args(argv)


# ---------------------------------------------------------------- 周划分

def week_key(day: date, mode: str) -> tuple[str, str, str]:
    """返回 (月目录, 周标签, 周内排序键)。"""
    month = f"{day.year:04d}-{day.month:02d}"
    if mode == "month":
        index = (day.day - 1) // 7 + 1
        return month, f"{day.month}-{index}", f"{index:02d}"
    iso_year, iso_week, _ = day.isocalendar()
    return month, f"W{iso_week:02d}", f"{iso_week:02d}"


def week_span(days: list[str]) -> str:
    first, last = to_date(days[0]), to_date(days[-1])
    return f"{first.month:02d}{first.day:02d}-{last.month:02d}{last.day:02d}"


# ---------------------------------------------------------------- 写盘

def style_header(sheet, headers, row: int = 1):
    sheet.append([label for _, label, _ in headers])
    for cell in sheet[row]:
        cell.fill = PatternFill("solid", fgColor=HEAD_FILL)
        cell.font = Font(color=WHITE, bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[row].height = 22
    for index, (_, _, width) in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = f"A{row + 1}"


def write_week(path: Path, province: str, label: str, rows: list[dict], daily: list[dict]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "分时明细"
    sheet.sheet_view.showGridLines = False
    style_header(sheet, DETAIL_HEADERS)
    for row in rows:
        sheet.append([row.get(key) for key, _, _ in DETAIL_HEADERS])
    sheet.auto_filter.ref = sheet.dimensions

    summary = workbook.create_sheet("每日汇总")
    summary.sheet_view.showGridLines = False
    style_header(summary, DAILY_HEADERS)
    for row in daily:
        summary.append([row.get(key) for key, _, _ in DAILY_HEADERS])

    info = workbook.create_sheet("说明")
    info.sheet_view.showGridLines = False
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 62
    pairs = [
        ("区域", province),
        ("周次", label),
        ("日期范围", f"{daily[0]['trade_date']} ~ {daily[-1]['trade_date']}" if daily else ""),
        ("覆盖天数", len(daily)),
        ("分时点数", len(rows)),
        ("价格单位", PRICE_UNIT),
        ("生成时间", now_stamp()),
        ("口径", "按自然日逐日请求，保留接口原生日内粒度；缺失值留空不按 0 处理。"),
        ("提示", "部分区域接口只返回实时价，日前价为空属接口口径。"),
    ]
    for index, (key, value) in enumerate(pairs, 1):
        info.cell(index, 1, key).font = Font(bold=True)
        info.cell(index, 1).fill = PatternFill("solid", fgColor=SUB_FILL)
        info.cell(index, 2, value)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_daily_book(path: Path, title: str, sheets: list[tuple[str, list[dict]]]):
    """月汇总 / 全期汇总：只放每日汇总，按周或按月分表。"""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets:
        sheet = workbook.create_sheet(name[:31])
        sheet.sheet_view.showGridLines = False
        style_header(sheet, DAILY_HEADERS)
        for row in rows:
            sheet.append([row.get(key) for key, _, _ in DAILY_HEADERS])
        sheet.auto_filter.ref = sheet.dimensions
    if not workbook.sheetnames:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


# ---------------------------------------------------------------- 主流程

def main(argv=None) -> int:
    args = parse_args(argv)
    raw_dir = args.raw_dir.resolve()
    out_dir = args.output_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    provinces = load_provinces()
    if args.only_provinces:
        wanted = set(args.only_provinces)
        provinces = [
            p for p in provinces
            if p["province"] in wanted or p["province_code"] in wanted or p["province_type"] in wanted
        ]
        if not provinces:
            raise SystemExit("--only-provinces 未匹配到任何区域")
    name_by_code = {p["province_code"]: p["province"] for p in provinces }
    wanted_codes = set(name_by_code)

    # 一次流式扫描，按 区域 → 周 分桶
    buckets: dict[tuple[str, str, str, str], list[dict]] = defaultdict(list)
    for row in iter_csv(raw_dir / DETAIL_FILE):
        code = row["province_code"]
        if code not in wanted_codes:
            continue
        day = to_date(row["trade_date"])
        month, label, order = week_key(day, args.week_mode)
        da = as_number(row["day_ahead_price"])
        rt = as_number(row["real_time_price"])
        buckets[(code, month, order, label)].append({
            "trade_date": row["trade_date"],
            "weekday": WEEKDAYS[day.weekday()],
            "time_slot": row["time_slot"],
            "day_ahead_price": da,
            "real_time_price": rt,
            "spread": None if (da is None or rt is None) else round(rt - da, 2),
        })
    if not buckets:
        raise SystemExit("数据仓里没有可导出的明细")

    state_path = out_dir / STATE_FILE
    state: dict[str, str] = {}
    if state_path.exists() and not args.full:
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
        except ValueError:
            state = {}
    new_state: dict[str, str] = {}

    written = skipped = 0
    per_province_daily: dict[str, list[dict]] = defaultdict(list)
    per_month_weeks: dict[tuple[str, str], list[tuple[str, list[dict]]]] = defaultdict(list)
    index_rows: list[dict] = []

    for (code, month, order, label), rows in sorted(buckets.items()):
        province = name_by_code[code]
        rows.sort(key=lambda r: (r["trade_date"], r["time_slot"]))

        by_day: dict[str, list[dict]] = defaultdict(list)
        for row in rows:
            by_day[row["trade_date"]].append(row)
        daily = []
        for trade_date, day_rows in sorted(by_day.items()):
            da = [r["day_ahead_price"] for r in day_rows if r["day_ahead_price"] is not None]
            rt = [r["real_time_price"] for r in day_rows if r["real_time_price"] is not None]
            da_avg, rt_avg = mean(da), mean(rt)
            daily.append({
                "trade_date": trade_date,
                "weekday": day_rows[0]["weekday"],
                "point_count": len(day_rows),
                "day_ahead_avg": round(da_avg, 2) if da_avg != "" else None,
                "real_time_avg": round(rt_avg, 2) if rt_avg != "" else None,
                "spread_avg": round(rt_avg - da_avg, 2) if da and rt else None,
                "day_ahead_min": round(min(da), 2) if da else None,
                "day_ahead_max": round(max(da), 2) if da else None,
                "real_time_min": round(min(rt), 2) if rt else None,
                "real_time_max": round(max(rt), 2) if rt else None,
                "negative_points": sum(v < 0 for v in rt),
                "zero_points": sum(v == 0 for v in rt),
            })

        days = [d["trade_date"] for d in daily]
        span = week_span(days)
        filename = f"{province}_{month}_{label}_{span}.xlsx"
        path = out_dir / province / month / filename
        key = f"{province}/{month}/{label}"

        # 指纹只看数据本身，避免"生成时间"变化导致每次都重写
        digest = hashlib.sha256(
            json.dumps(rows, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
        ).hexdigest()[:16]
        new_state[key] = digest

        if state.get(key) == digest and path.exists():
            skipped += 1
        else:
            # 周边界或命名规则变化时，先清掉同一周的旧文件
            for stale in (out_dir / province / month).glob(f"{province}_{month}_{label}_*.xlsx"):
                if stale != path:
                    stale.unlink()
            write_week(path, province, label, rows, daily)
            written += 1

        per_province_daily[province].extend(daily)
        per_month_weeks[(province, month)].append((label, daily))
        index_rows.append({
            "province": province,
            "month": month,
            "label": label,
            "span": span,
            "days": len(daily),
            "points": len(rows),
            "file": str(path.relative_to(out_dir)),
        })

    if not args.no_summary:
        for (province, month), weeks in sorted(per_month_weeks.items()):
            write_daily_book(
                out_dir / province / month / f"{province}_{month}_月汇总.xlsx",
                f"{province} {month}",
                sorted(weeks, key=lambda w: w[0]),
            )
        for province, daily in sorted(per_province_daily.items()):
            by_month: dict[str, list[dict]] = defaultdict(list)
            for row in daily:
                by_month[row["trade_date"][:7]].append(row)
            write_daily_book(
                out_dir / province / f"{province}_全期汇总.xlsx",
                province,
                [(m, sorted(rows, key=lambda r: r["trade_date"])) for m, rows in sorted(by_month.items())],
            )

    # 索引表：一眼看清每个区域各周的文件与覆盖情况
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "文件索引"
    sheet.sheet_view.showGridLines = False
    headers = [("province", "区域", 12), ("month", "月份", 11), ("label", "周次", 9),
               ("span", "日期范围", 13), ("days", "天数", 8), ("points", "分时点数", 11),
               ("file", "相对路径", 56)]
    style_header(sheet, headers)
    for row in index_rows:
        sheet.append([row[key] for key, _, _ in headers])
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(out_dir / "_索引.xlsx")

    state_path.write_text(json.dumps(new_state, ensure_ascii=False, indent=1), encoding="utf-8")

    meta = json.loads((raw_dir / METADATA_FILE).read_text(encoding="utf-8"))
    print(
        f"分层导出完成｜区域 {len(per_province_daily)}｜周文件 {len(index_rows)}"
        f"（新写 {written}，未变跳过 {skipped}）｜周划分 {args.week_mode}"
    )
    print(f"输出目录 {out_dir.relative_to(config.ROOT)}，数据区间 {meta['start_date']} ~ {meta['end_date']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
