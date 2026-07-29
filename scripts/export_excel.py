#!/usr/bin/env python3
"""把 CSV 数据仓导出为一个汇总工作簿 + 若干按月拆分的明细工作簿。

按月拆分是为了避免几十万行挤进同一个 Excel 造成打开缓慢或内存不足。
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

import config  # noqa: E402
from common import (  # noqa: E402
    DAILY_FILE,
    DETAIL_FILE,
    METADATA_FILE,
    PROVINCE_SUMMARY_FILE,
    QUALITY_FILE,
    iter_csv,
    read_csv,
)

BLUE = "1D4ED8"
LIGHT_BLUE = "DBEAFE"
WHITE = "FFFFFF"
TEXT = "172033"

STATUS_LABELS = {
    "available": "有数据",
    "empty": "接口无数据",
    "failed": "请求失败",
    "missing": "尚未采集",
}

PROVINCE_HEADERS = [
    ("province", "区域"),
    ("province_code", "区域代码"),
    ("coverage", "覆盖率"),
    ("requested_days", "请求天数"),
    ("available_days", "有数据天数"),
    ("empty_days", "空数据天数"),
    ("failed_days", "失败天数"),
    ("missing_days", "未采集天数"),
    ("dominant_points_per_day", "主流日点数"),
    ("irregular_point_days", "异常点数天数"),
    ("point_count", "明细点数"),
    ("day_ahead_avg", "日前均价"),
    ("real_time_avg", "实时均价"),
    ("spread_avg", "实时-日前价差"),
    ("day_ahead_min", "日前最低"),
    ("day_ahead_max", "日前最高"),
    ("real_time_min", "实时最低"),
    ("real_time_max", "实时最高"),
    ("negative_realtime_points", "实时负价点数"),
    ("zero_realtime_points", "实时零价点数"),
    ("first_date", "最早日期"),
    ("last_date", "最晚日期"),
]
DAILY_HEADERS = [
    ("province", "区域"),
    ("province_code", "区域代码"),
    ("trade_date", "交易日期"),
    ("point_count", "日点数"),
    ("day_ahead_avg", "日前均价"),
    ("real_time_avg", "实时均价"),
    ("spread_avg", "实时-日前价差"),
    ("day_ahead_min", "日前最低"),
    ("day_ahead_max", "日前最高"),
    ("real_time_min", "实时最低"),
    ("real_time_max", "实时最高"),
]
DETAIL_HEADERS = [
    ("province", "区域"),
    ("province_code", "区域代码"),
    ("province_type", "区域类型"),
    ("trade_date", "交易日期"),
    ("time_slot", "时点"),
    ("day_ahead_price", "日前价格（元/MWh）"),
    ("real_time_price", "实时价格（元/MWh）"),
    ("unit", "单位"),
    ("collected_at", "采集时间"),
]


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="把数据仓导出为汇总及月度 Excel")
    parser.add_argument("--raw-dir", type=Path, default=config.RAW_DIR)
    parser.add_argument("--output-dir", type=Path, default=config.EXCEL_DIR)
    parser.add_argument("--no-detail", action="store_true", help="只导出汇总工作簿")
    return parser.parse_args(argv)


def coerce(value: str):
    if value == "" or value is None:
        return None
    try:
        return float(value) if "." in str(value) else int(value)
    except (TypeError, ValueError):
        return value


def add_csv_sheet(workbook: Workbook, title: str, rows: list[dict], headers: list[tuple[str, str]]):
    worksheet = workbook.create_sheet(title)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.append([label for _, label in headers])
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        worksheet.append([coerce(row.get(key, "")) for key, _ in headers])
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 28
    for index, (key, label) in enumerate(headers, 1):
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = min(38, max(12, len(label) * 2 + 4))
        if key == "coverage":
            for cell in worksheet[letter][1:]:
                cell.number_format = "0.0%"
    return worksheet


def save_summary(raw_dir: Path, output_path: Path, metadata: dict):
    workbook = Workbook()
    overview = workbook.active
    overview.title = "使用说明"
    overview.sheet_view.showGridLines = False
    overview.merge_cells("A1:H2")
    overview["A1"] = "全国现货电价分时数据"
    overview["A1"].fill = PatternFill("solid", fgColor=BLUE)
    overview["A1"].font = Font(color=WHITE, bold=True, size=20)
    overview["A1"].alignment = Alignment(vertical="center")
    overview["A4"] = "指标"
    overview["B4"] = "数值"
    for cell in overview[4][:2]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(color=TEXT, bold=True)
    values = [
        ("价格区域", metadata.get("region_count")),
        ("采集起始日", metadata.get("start_date")),
        ("采集结束日", metadata.get("end_date")),
        ("请求天数", metadata.get("requested_days")),
        ("已采集区域日", metadata.get("available_region_days")),
        ("目标区域日", metadata.get("expected_region_days")),
        ("覆盖率", f"{metadata.get('coverage', 0):.1%}"),
        ("分时明细点数", metadata.get("detail_rows")),
        ("更新时间", metadata.get("updated_at")),
    ]
    for row_index, pair in enumerate(values, 5):
        overview.cell(row_index, 1, pair[0])
        overview.cell(row_index, 2, pair[1])
    overview["D4"] = "数据口径"
    overview["D4"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    overview["D4"].font = Font(color=TEXT, bold=True)
    notes = [
        "每个价格区域按自然日请求，startDate=endDate，避免接口返回区间平均值。",
        "保留接口返回的原生 24/48/96 等日内点数，不强制补齐。",
        "价格单位为元/MWh；缺失值不按 0 处理。",
        "明细按月份拆分为独立 Excel，避免单文件过大。",
        "文件中不保存 Authorization 令牌。",
        "覆盖率 = 有数据区域日 ÷ 目标区域日。",
    ]
    for index, note in enumerate(notes, 5):
        overview.cell(index, 4, note)
    overview.column_dimensions["A"].width = 18
    overview.column_dimensions["B"].width = 26
    overview.column_dimensions["C"].width = 4
    overview.column_dimensions["D"].width = 68

    add_csv_sheet(workbook, "区域汇总", read_csv(raw_dir / PROVINCE_SUMMARY_FILE), PROVINCE_HEADERS)
    add_csv_sheet(workbook, "每日汇总", read_csv(raw_dir / DAILY_FILE), DAILY_HEADERS)
    add_csv_sheet(
        workbook,
        "区域代码",
        read_csv(raw_dir / "province_codes.csv"),
        [("province", "区域"), ("province_code", "区域代码"), ("province_type", "区域类型")],
    )
    quality = [
        {**row, "status": STATUS_LABELS.get(row["status"], row["status"])}
        for row in read_csv(raw_dir / QUALITY_FILE)
    ]
    add_csv_sheet(
        workbook,
        "数据质量",
        quality,
        [
            ("province", "区域"),
            ("province_code", "区域代码"),
            ("trade_date", "交易日期"),
            ("status", "状态"),
            ("point_count", "返回点数"),
            ("error", "错误信息"),
            ("collected_at", "采集时间"),
        ],
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def save_month(rows: list[dict], month: str, output_path: Path):
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(f"明细_{month[:4]}-{month[4:]}")
    worksheet.append([label for _, label in DETAIL_HEADERS])
    for row in rows:
        worksheet.append([coerce(row.get(key, "")) for key, _ in DETAIL_HEADERS])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main(argv=None) -> int:
    args = parse_args(argv)
    raw_dir = args.raw_dir.resolve()
    output_dir = args.output_dir.resolve()
    metadata = json.loads((raw_dir / METADATA_FILE).read_text(encoding="utf-8"))
    span = f"{metadata['start_date'].replace('-', '')}-{metadata['end_date'].replace('-', '')}"
    summary_path = output_dir / f"全国现货电价_汇总与质量_{span}.xlsx"
    save_summary(raw_dir, summary_path, metadata)
    print(f"已生成 {summary_path.relative_to(config.ROOT)}")

    if args.no_detail:
        return 0

    by_month: dict[str, list[dict]] = defaultdict(list)
    for row in iter_csv(raw_dir / DETAIL_FILE):
        by_month[row["trade_date"][:7].replace("-", "")].append(row)
    for month, rows in sorted(by_month.items()):
        output_path = output_dir / f"全国现货电价_明细_{month}.xlsx"
        save_month(rows, month, output_path)
        print(f"已生成 {output_path.relative_to(config.ROOT)}：{len(rows)} 行")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
