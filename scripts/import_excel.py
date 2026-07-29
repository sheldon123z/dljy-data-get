#!/usr/bin/env python3
"""把历史 Excel 明细反向导入 CSV 数据仓，让旧成果可以被续采与再导出。

早期采集只留下了 Excel 合集，没有 CSV 仓库。本脚本读取每个工作簿里的
“明细_YYYY-MM”工作表，还原成 electricity_price_detail.csv，再由
common.build_outputs 重算质量表与各级汇总。
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook  # noqa: E402

import config  # noqa: E402
from common import (  # noqa: E402
    DETAIL_FIELDS,
    DETAIL_FILE,
    PRICE_UNIT,
    build_outputs,
    load_provinces,
    read_csv,
    to_date,
    write_csv,
)

# 明细工作表的中文表头 -> 仓库字段
HEADER_MAP = {
    "区域": "province",
    "区域代码": "province_code",
    "区域类型": "province_type",
    "交易日期": "trade_date",
    "时点": "time_slot",
    "日前价格（元/MWh）": "day_ahead_price",
    "实时价格（元/MWh）": "real_time_price",
    "单位": "unit",
    "采集时间": "collected_at",
}


def parse_args():
    parser = argparse.ArgumentParser(description="从历史 Excel 合集导入 CSV 数据仓")
    parser.add_argument(
        "--source",
        type=Path,
        nargs="+",
        help="Excel 文件或目录，默认扫描 data/archive/ 下所有 *明细*.xlsx",
    )
    parser.add_argument("--raw-dir", type=Path, default=config.RAW_DIR)
    parser.add_argument("--province-file", type=Path)
    parser.add_argument(
        "--replace",
        action="store_true",
        help="丢弃仓库中的已有明细，只保留本次导入结果",
    )
    return parser.parse_args()


def discover_workbooks(sources: list[Path] | None) -> list[Path]:
    if not sources:
        # 默认只看归档目录，避免把自己刚导出的 Excel 再读一遍
        sources = [config.ARCHIVE_DIR if config.ARCHIVE_DIR.exists() else config.DATA_DIR]
    files: list[Path] = []
    for item in sources:
        item = item.expanduser()
        if item.is_dir():
            files.extend(sorted(item.rglob("*.xlsx")))
        elif item.exists():
            files.append(item)
    return [
        path
        for path in dict.fromkeys(files)
        if not path.name.startswith((".", "~$")) and "明细" in path.name
    ]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_detail_sheets(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    try:
        for name in workbook.sheetnames:
            if not name.startswith("明细"):
                continue
            sheet = workbook[name]
            stream = sheet.iter_rows(values_only=True)
            try:
                header = next(stream)
            except StopIteration:
                continue
            columns = [HEADER_MAP.get(clean(cell)) for cell in header]
            if "trade_date" not in columns or "time_slot" not in columns:
                continue
            for raw in stream:
                record = {
                    field: clean(value)
                    for field, value in zip(columns, raw)
                    if field
                }
                if not record.get("province_code") or not record.get("trade_date"):
                    continue
                record["trade_date"] = to_date(record["trade_date"]).isoformat()
                record.setdefault("unit", PRICE_UNIT)
                record.setdefault("collected_at", "")
                rows.append(record)
    finally:
        workbook.close()
    return rows


def main() -> int:
    args = parse_args()
    config.ensure_dirs()
    provinces = load_provinces(args.province_file)
    type_by_code = {row["province_code"]: row["province_type"] for row in provinces}
    name_by_code = {row["province_code"]: row["province"] for row in provinces}

    workbooks = discover_workbooks(args.source)
    if not workbooks:
        raise SystemExit("未找到可导入的 Excel 明细文件")

    raw_dir = args.raw_dir.resolve()
    raw_dir.mkdir(parents=True, exist_ok=True)
    detail_path = raw_dir / DETAIL_FILE

    merged: dict[tuple[str, str, str], dict] = {}
    if not args.replace:
        for row in read_csv(detail_path):
            merged[(row["province_code"], row["trade_date"], row["time_slot"])] = row

    for path in workbooks:
        rows = read_detail_sheets(path)
        for row in rows:
            code = row["province_code"]
            row["province"] = row.get("province") or name_by_code.get(code, "")
            row["province_type"] = row.get("province_type") or type_by_code.get(code, "")
            merged[(code, row["trade_date"], row["time_slot"])] = row
        print(f"读取 {path.name}：{len(rows)} 行", flush=True)

    if not merged:
        raise SystemExit("Excel 中没有解析到任何明细行")

    details = sorted(
        merged.values(),
        key=lambda row: (row["province_code"], row["trade_date"], row["time_slot"]),
    )
    write_csv(detail_path, DETAIL_FIELDS, details)

    dates = sorted({row["trade_date"] for row in details})
    metadata = build_outputs(
        raw_dir,
        provinces,
        dates[0],
        dates[-1],
        collection_status="imported_from_excel",
    )
    print(
        f"导入完成：{metadata['detail_rows']} 个分时点，"
        f"{metadata['available_region_days']}/{metadata['expected_region_days']} 个有数据区域日"
        f"（覆盖率 {metadata['coverage']:.1%}），区间 {metadata['start_date']} ~ {metadata['end_date']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
